import openpyxl

# 读取原始 Excel 文件
wb = openpyxl.load_workbook('desensitizer-spring-boot/src/main/resources/test-data/赛题4-支持敏感信息脱敏的通用工具-测试数据v0.3.xlsx')
ws = wb.active

# 在第2行（表头行）添加新列名
ws.cell(row=2, column=7, value='姓名_脱敏')
ws.cell(row=2, column=8, value='手机号_脱敏')
ws.cell(row=2, column=9, value='身份证号_脱敏')
ws.cell(row=2, column=10, value='银行卡号_脱敏')
ws.cell(row=2, column=11, value='地址_脱敏')
ws.cell(row=2, column=12, value='国家_脱敏')

# 遍历数据行（从第3行开始）
for row in range(3, ws.max_row + 1):
    # 姓名脱敏
    name = str(ws.cell(row=row, column=1).value)
    if len(name) == 2:
        name_desensitized = name[0] + '*'
    elif len(name) == 3:
        name_desensitized = name[0] + '*' + name[-1]
    elif len(name) >= 4:
        name_desensitized = name[0] + '**' + name[-1]
    else:
        name_desensitized = name
    ws.cell(row=row, column=7, value=name_desensitized)
    
    # 手机号脱敏
    phone = str(ws.cell(row=row, column=2).value).strip()
    if len(phone) >= 11:
        phone_desensitized = phone[:3] + '****' + phone[-4:]
    else:
        phone_desensitized = phone
    ws.cell(row=row, column=8, value=phone_desensitized)
    
    # 身份证号脱敏
    id_card = str(ws.cell(row=row, column=3).value).strip()
    if len(id_card) >= 18:
        id_card_desensitized = id_card[:6] + '********' + id_card[-4:]
    else:
        id_card_desensitized = id_card
    ws.cell(row=row, column=9, value=id_card_desensitized)
    
    # 银行卡号脱敏
    bank_card = str(ws.cell(row=row, column=4).value).strip()
    if len(bank_card) >= 16:
        bank_card_desensitized = bank_card[:6] + '********' + bank_card[-4:]
    else:
        bank_card_desensitized = bank_card
    ws.cell(row=row, column=10, value=bank_card_desensitized)
    
    # 地址脱敏（与 Java AddressDesensitizer 保持一致）
    address = str(ws.cell(row=row, column=5).value).strip()
    
    # 辅助函数：检查是否包含中文字符
    def contains_chinese(text):
        for c in text:
            if '\u4e00' <= c <= '\u9fa5':
                return True
        return False
    
    # 只支持中国地址脱敏
    if not contains_chinese(address):
        addr_desensitized = address
    # 中文地址处理：优先匹配行政区域关键词
    elif '区' in address or '县' in address:
        idx_district = address.rfind('区')
        idx_county = address.rfind('县')
        idx = max(idx_district, idx_county)
        if idx > 0:
            addr_desensitized = address[:idx+1] + '***'
        else:
            addr_desensitized = address
    elif '市' in address:
        idx = address.find('市')
        if idx > 0:
            addr_desensitized = address[:idx+1] + '***'
        else:
            addr_desensitized = address
    elif '省' in address:
        idx = address.find('省')
        if idx > 0:
            addr_desensitized = address[:idx+1] + '***'
        else:
            addr_desensitized = address
    # 默认处理：保留前半部分，脱敏后半部分
    else:
        half_length = len(address) // 2
        if half_length >= 2:
            addr_desensitized = address[:half_length] + '***'
        elif len(address) > 4:
            addr_desensitized = address[:4] + '***'
        else:
            addr_desensitized = address
    ws.cell(row=row, column=11, value=addr_desensitized)
    
    # 国家脱敏（不脱敏）
    country = ws.cell(row=row, column=6).value
    ws.cell(row=row, column=12, value=str(country) if country else '')

wb.save('desensitizer-spring-boot/src/main/resources/test-data/赛题4-支持敏感信息脱敏的通用工具-测试数据v0.3.xlsx')
print(f'已处理 {ws.max_row - 2} 行数据，添加了预期脱敏结果列！')
